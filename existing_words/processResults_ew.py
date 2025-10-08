"""
Process results from existing word experiments.

This script processes the results of existing word experiments by:
1. Finding experiments in outputs/ directories containing .xlsx files with 'number' or 'weighted_sum' in names
2. Extracting dataset files from config.yaml for each experiment
3. Comparing 'exist' predictions with 'answer' ground truth
4. Generating results with accuracy calculations

Usage:
    python processResults_ew.py all          # Process all experiments
    python processResults_ew.py <experiment> # Process specific experiment
"""

import pandas as pd
import os
import glob
import yaml
import sys
from datetime import datetime


def load_config():
    """Load the configuration file."""
    with open('config.yaml', 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)


def find_experiments():
    """Find all experiments that have output files with 'number' or 'weighted_sum' in their names."""
    experiments = []
    outputs_dir = 'outputs'
    
    if not os.path.exists(outputs_dir):
        print(f"Error: {outputs_dir} directory not found")
        return experiments
    
    # Get all experiment directories
    exp_dirs = [d for d in os.listdir(outputs_dir) if os.path.isdir(os.path.join(outputs_dir, d))]
    
    for exp_dir in exp_dirs:
        exp_path = os.path.join(outputs_dir, exp_dir)
        # Look for .xlsx files containing 'number' or 'weighted_sum'
        xlsx_files = glob.glob(os.path.join(exp_path, '*.xlsx'))
        
        for xlsx_file in xlsx_files:
            filename = os.path.basename(xlsx_file)
            if 'number' in filename or 'weighted_sum' in filename:
                experiments.append({
                    'name': exp_dir,
                    'output_file': xlsx_file
                })
                break  # Only need one matching file per experiment
    
    return experiments


def extract_output_data(output_file):
    """Extract 'word' and 'exist' columns from output file."""
    try:
        df = pd.read_excel(output_file)
        if 'word' not in df.columns or 'exist' not in df.columns:
            print(f"Error: Required columns not found in {output_file}")
            return None
        return df[['word', 'exist']].copy()
    except Exception as e:
        print(f"Error reading {output_file}: {e}")
        return None


def extract_dataset_data(dataset_path):
    """Extract 'word' and 'answer' columns from dataset file."""
    try:
        full_path = os.path.join('data', dataset_path)
        df = pd.read_excel(full_path)
        
        # Handle different possible column names for word
        word_col = None
        for col in df.columns:
            if col.lower() in ['word', 'Word']:
                word_col = col
                break
        
        if word_col is None or 'answer' not in df.columns:
            print(f"Error: Required columns not found in {full_path}")
            print(f"Available columns: {df.columns.tolist()}")
            return None
            
        return df[[word_col, 'answer']].rename(columns={word_col: 'word'}).copy()
    except Exception as e:
        print(f"Error reading {dataset_path}: {e}")
        return None


def compare_results(output_data, dataset_data):
    """Compare exist predictions with expected answers."""
    # Merge on word column
    merged = pd.merge(output_data, dataset_data, on='word', how='inner')
    
    if len(merged) == 0:
        print("Error: No matching words found between output and dataset")
        return None
    
    # Compare exist and answer to determine correctness
    merged['correct'] = (merged['exist'] == merged['answer']).astype(int)
    
    # Rename columns as specified
    merged = merged.rename(columns={
        'exist': 'obtained_answer',
        'answer': 'expected_answer'
    })
    
    return merged


def calculate_accuracy(results_df):
    """Calculate accuracy from results dataframe."""
    total_words = len(results_df)
    correct_predictions = results_df['correct'].sum()
    accuracy = correct_predictions / total_words if total_words > 0 else 0
    return accuracy, correct_predictions, total_words


def save_results(results_df, experiment_name, accuracy):
    """Save results to Excel file with accuracy included and formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Create processResults directory if it doesn't exist
    results_dir = os.path.join('processResults', experiment_name)
    os.makedirs(results_dir, exist_ok=True)
    
    # Generate timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{experiment_name}_results_{timestamp}.xlsx"
    filepath = os.path.join(results_dir, filename)
    
    # Create a new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    
    # Add the dataframe data to the worksheet
    for r in dataframe_to_rows(results_df, index=False, header=True):
        ws.append(r)
    
    # Add accuracy in a separate row at the end
    ws.append([])  # Empty row
    ws.append(['Accuracy:', accuracy])
    
    # Define fill colors
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    red_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")   # Light red
    
    # Find the 'correct' column index
    correct_col_idx = None
    for col_idx, cell in enumerate(ws[1], 1):  # First row contains headers
        if cell.value == 'correct':
            correct_col_idx = col_idx
            break
    
    # Apply conditional formatting to the 'correct' column
    if correct_col_idx:
        for row_idx in range(2, len(results_df) + 2):  # Skip header row
            cell = ws.cell(row=row_idx, column=correct_col_idx)
            if cell.value == 1:
                cell.fill = green_fill
            elif cell.value == 0:
                cell.fill = red_fill
    
    # Save the workbook
    wb.save(filepath)
    return filepath


def process_experiment(experiment_name, config):
    """Process a single experiment."""
    print(f"\nProcessing experiment: {experiment_name}")
    
    # Find the output file for this experiment
    experiments = find_experiments()
    experiment_data = None
    for exp in experiments:
        if exp['name'] == experiment_name:
            experiment_data = exp
            break
    
    if experiment_data is None:
        print(f"Error: No output file found for experiment {experiment_name}")
        return False
    
    # Get dataset path from config
    if experiment_name not in config['experiments']:
        print(f"Error: Experiment {experiment_name} not found in config.yaml")
        return False
    
    dataset_path = config['experiments'][experiment_name]['dataset_path']
    
    # Extract data from both files
    output_data = extract_output_data(experiment_data['output_file'])
    if output_data is None:
        return False
    
    dataset_data = extract_dataset_data(dataset_path)
    if dataset_data is None:
        return False
    
    # Compare results
    results = compare_results(output_data, dataset_data)
    if results is None:
        return False
    
    # Calculate accuracy
    accuracy, correct, total = calculate_accuracy(results)
    
    # Save results
    filepath = save_results(results, experiment_name, accuracy)
    
    # Print results
    print(f"Experiment: {experiment_name}")
    print(f"Accuracy: {accuracy:.4f} ({correct}/{total})")
    print(f"Results saved to: {filepath}")
    
    return True


def main():
    """Main function to process experiments."""
    if len(sys.argv) < 2:
        print("Usage: python processResults_ew.py <experiment_name|all>")
        sys.exit(1)
    
    target = sys.argv[1]
    
    # Load configuration
    config = load_config()
    
    if target.lower() == 'all':
        # Process all experiments
        experiments = find_experiments()
        if not experiments:
            print("No experiments found")
            return
        
        successful = 0
        for exp in experiments:
            if process_experiment(exp['name'], config):
                successful += 1
        
        print(f"\nProcessed {successful}/{len(experiments)} experiments successfully")
    else:
        # Process specific experiment
        if not process_experiment(target, config):
            print(f"Failed to process experiment: {target}")
            sys.exit(1)


if __name__ == "__main__":
    main()

